"""
Export routes - Multi-format export functionality
"""
import io
import csv
import json
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from models.listing import Listing
from utils.auth import token_required
from utils.audit import log_action

export_bp = Blueprint('export', __name__)


@export_bp.route('/text', methods=['POST'])
@token_required
def export_text(current_user):
    """Export listings as tab-delimited text"""
    listing_ids = request.json.get('listing_ids', [])
    
    if listing_ids:
        listings = Listing.query.filter(
            Listing.id.in_(listing_ids),
            Listing.user_id == current_user.id
        ).all()
    else:
        listings = Listing.query.filter_by(user_id=current_user.id).all()
    
    if not listings:
        return jsonify({'error': 'No listings found'}), 404
    
    # Generate text output
    output = io.StringIO()
    
    # Header
    output.write("TITLE\tPRICE\tCONDITION\tDESCRIPTION\tCATEGORY\tOFFER SHIPPING\n")
    
    # Data rows
    for listing in listings:
        output.write(f"{listing.title}\t")
        output.write(f"{listing.price}\t")
        output.write(f"{listing.condition}\t")
        output.write(f"{listing.description or ''}\t")
        output.write(f"{listing.category or ''}\t")
        output.write(f"{listing.offer_shipping or 'No'}\n")
    
    log_action(current_user.id, 'export_text', 'listing', None, 200, 
               metadata={'count': len(listings)})
    
    # Return as file
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=f'marketplace-listings-{datetime.now().strftime("%Y%m%d-%H%M%S")}.txt'
    )


@export_bp.route('/csv', methods=['POST'])
@token_required
def export_csv(current_user):
    """Export listings as CSV"""
    listing_ids = request.json.get('listing_ids', [])
    
    if listing_ids:
        listings = Listing.query.filter(
            Listing.id.in_(listing_ids),
            Listing.user_id == current_user.id
        ).all()
    else:
        listings = Listing.query.filter_by(user_id=current_user.id).all()
    
    if not listings:
        return jsonify({'error': 'No listings found'}), 404
    
    # Generate CSV output
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['TITLE', 'PRICE', 'CONDITION', 'DESCRIPTION', 'CATEGORY', 'OFFER SHIPPING'])
    
    # Data rows
    for listing in listings:
        writer.writerow([
            listing.title,
            float(listing.price) if listing.price else '',
            listing.condition,
            listing.description or '',
            listing.category or '',
            listing.offer_shipping or 'No'
        ])
    
    log_action(current_user.id, 'export_csv', 'listing', None, 200,
               metadata={'count': len(listings)})
    
    # Return as file
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'marketplace-listings-{datetime.now().strftime("%Y%m%d-%H%M%S")}.csv'
    )


@export_bp.route('/json', methods=['POST'])
@token_required
def export_json(current_user):
    """Export listings as JSON"""
    listing_ids = request.json.get('listing_ids', [])
    
    if listing_ids:
        listings = Listing.query.filter(
            Listing.id.in_(listing_ids),
            Listing.user_id == current_user.id
        ).all()
    else:
        listings = Listing.query.filter_by(user_id=current_user.id).all()
    
    if not listings:
        return jsonify({'error': 'No listings found'}), 404
    
    # Convert to Facebook format
    data = [listing.to_facebook_format() for listing in listings]
    
    log_action(current_user.id, 'export_json', 'listing', None, 200,
               metadata={'count': len(listings)})
    
    # Return as file
    json_str = json.dumps(data, indent=2)
    return send_file(
        io.BytesIO(json_str.encode('utf-8')),
        mimetype='application/json',
        as_attachment=True,
        download_name=f'marketplace-listings-{datetime.now().strftime("%Y%m%d-%H%M%S")}.json'
    )


@export_bp.route('/xlsx', methods=['POST'])
@token_required
def export_xlsx(current_user):
    """Export listings as Excel (XLSX)"""
    listing_ids = request.json.get('listing_ids', [])

    if listing_ids:
        listings = Listing.query.filter(
            Listing.id.in_(listing_ids),
            Listing.user_id == current_user.id
        ).all()
    else:
        listings = Listing.query.filter_by(user_id=current_user.id).all()

    if not listings:
        return jsonify({'error': 'No listings found'}), 404

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Marketplace Listings"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Headers
    headers = ['TITLE', 'PRICE', 'CONDITION', 'DESCRIPTION', 'CATEGORY', 'OFFER SHIPPING']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_num, listing in enumerate(listings, 2):
        ws.cell(row=row_num, column=1, value=listing.title)
        ws.cell(row=row_num, column=2, value=float(listing.price) if listing.price else '')
        ws.cell(row=row_num, column=3, value=listing.condition)
        ws.cell(row=row_num, column=4, value=listing.description or '')
        ws.cell(row=row_num, column=5, value=listing.category or '')
        ws.cell(row=row_num, column=6, value=listing.offer_shipping or 'No')

    # Adjust column widths
    ws.column_dimensions['A'].width = 50  # TITLE
    ws.column_dimensions['B'].width = 12  # PRICE
    ws.column_dimensions['C'].width = 20  # CONDITION
    ws.column_dimensions['D'].width = 60  # DESCRIPTION
    ws.column_dimensions['E'].width = 20  # CATEGORY
    ws.column_dimensions['F'].width = 15  # OFFER SHIPPING

    log_action(current_user.id, 'export_xlsx', 'listing', None, 200,
               metadata={'count': len(listings)})

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'marketplace-listings-{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx'
    )


@export_bp.route('/sql', methods=['POST'])
@token_required
def export_sql(current_user):
    """Export listings as SQL INSERT statements"""
    listing_ids = request.json.get('listing_ids', [])

    if listing_ids:
        listings = Listing.query.filter(
            Listing.id.in_(listing_ids),
            Listing.user_id == current_user.id
        ).all()
    else:
        listings = Listing.query.filter_by(user_id=current_user.id).all()

    if not listings:
        return jsonify({'error': 'No listings found'}), 404

    # Generate SQL output
    output = io.StringIO()

    # Table creation
    output.write("-- Marketplace Listings Export\n")
    output.write(f"-- Generated: {datetime.now().isoformat()}\n")
    output.write(f"-- Total listings: {len(listings)}\n\n")

    output.write("CREATE TABLE IF NOT EXISTS marketplace_listings (\n")
    output.write("    id VARCHAR(36) PRIMARY KEY,\n")
    output.write("    title VARCHAR(150) NOT NULL,\n")
    output.write("    price DECIMAL(10, 2) NOT NULL,\n")
    output.write("    condition VARCHAR(50) NOT NULL,\n")
    output.write("    description TEXT,\n")
    output.write("    category VARCHAR(100),\n")
    output.write("    offer_shipping VARCHAR(3),\n")
    output.write("    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP\n")
    output.write(");\n\n")

    # INSERT statements
    for listing in listings:
        title = listing.title.replace("'", "''")
        description = (listing.description or '').replace("'", "''")
        category = (listing.category or '').replace("'", "''")

        output.write(f"INSERT INTO marketplace_listings ")
        output.write(f"(id, title, price, condition, description, category, offer_shipping) VALUES (\n")
        output.write(f"    '{listing.id}',\n")
        output.write(f"    '{title}',\n")
        output.write(f"    {listing.price},\n")
        output.write(f"    '{listing.condition}',\n")
        output.write(f"    '{description}',\n")
        output.write(f"    '{category}',\n")
        output.write(f"    '{listing.offer_shipping or 'No'}'\n")
        output.write(f");\n\n")

    log_action(current_user.id, 'export_sql', 'listing', None, 200,
               metadata={'count': len(listings)})

    # Return as file
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=f'marketplace-listings-{datetime.now().strftime("%Y%m%d-%H%M%S")}.sql'
    )

